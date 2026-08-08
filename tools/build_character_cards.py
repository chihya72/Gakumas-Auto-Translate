#!/usr/bin/env python
"""从 csv_data 的人工译文统计生成角色卡草稿 tools/vendor/character-cards.json。

只做机械统计（自称、对他人的称呼、出场量），不调 API——统计结果是确定的，
每次跑出来一样；让模型归纳反而每次不同，一致性更差。

`speech`（语气/口癖规则）留空，必须人工填写。规则要可执行：
  ✗ "她说话很可爱"
  ✓ "句尾「〜だよね」译作「……对吧」，不要译成「……呢」"

用法：
    python tools/build_character_cards.py            # 生成草稿，保留已有人工字段
    python tools/build_character_cards.py --top 20   # 改角色数量阈值
"""
import argparse
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
CSV_DATA = ROOT / "csv_data"
NAME_DICT = ROOT / "name_dictionary.json"
OUT = ROOT / "tools" / "vendor" / "character-cards.json"

# 与 tm.ts 的 AI_TRANSLATOR_RE 保持一致
AI_TRANSLATOR_RE = re.compile(
    r"(gpt|deepseek|claude|gemini|glm|qwen|kimi|minimax|grok|o1|o3|llama"
    r"|mistral|sakura|ernie|hunyuan)", re.I)

FIRST_PERSON = ["わたくし", "あたし", "わたし", "私", "ボク", "ぼく", "僕",
                "オレ", "おれ", "俺", "うち"]
HONORIFICS = ["先輩", "せんぱい", "さん", "ちゃん", "くん", "君", "先生", "様", "さま"]
# 不建卡：玩家、旁白、训练员、学园长、元数据行、群众
SKIP_NAMES = {"", "{user}", "__narration__", "info", "译者", "一同", "モブ",
              "学園長", "トレーナー", "ダンストレーナー", "ボーカルトレーナー",
              "ビジュアルトレーナー"}
# 不建卡但仍要统计"别人怎么称呼他"——制作人的称呼是最影响连贯的一项
ADDRESS_EXTRA = ["プロデューサー"]


def is_human(translator):
    return bool(translator) and not AI_TRANSLATOR_RE.search(translator)


def load_rows():
    """读 csv_data 全部人工译文，返回 [(name, text, trans), ...]。"""
    rows = []
    files = 0
    for path in sorted(CSV_DATA.glob("*.csv")):
        try:
            with path.open(encoding="utf-8", newline="") as f:
                data = list(csv.DictReader(f))
        except (OSError, csv.Error) as e:
            print(f"!! 跳过无法读取的文件 {path.name}: {e}")
            continue
        translator = next((r.get("name", "") for r in data
                           if r.get("id") == "译者"), "")
        if not is_human(translator):
            continue
        files += 1
        for r in data:
            if r.get("id") in ("info", "译者"):
                continue
            name, text = r.get("name") or "", r.get("text") or ""
            if text:
                rows.append((name, text, r.get("trans") or ""))
    print(f"读入人工译文 {files} 个文件，{len(rows)} 行")
    return rows


def pick_cast(rows, top):
    counts = Counter(name for name, _, _ in rows if name not in SKIP_NAMES)
    cast = [n for n, _ in counts.most_common(top)]
    total = sum(counts.values())
    covered = sum(counts[n] for n in cast)
    pct = 100 * covered / total if total else 0
    print(f"选入 {len(cast)} 个角色，覆盖 {covered}/{total} 行（{pct:.1f}%）")
    return cast, counts


def first_person(lines):
    """按出现次数排出该角色的自称。较长的形式先匹配，避免 わたし 吃掉 あたし。"""
    hits = Counter()
    for text in lines:
        rest = text
        for token in FIRST_PERSON:
            n = rest.count(token)
            if n:
                hits[token] += n
                rest = rest.replace(token, "")
    return hits


def address_forms(pairs, cast, speaker):
    """统计该角色称呼其他角色时用的敬称后缀，并从对齐的译文里提取实际中文译法。

    pairs: [(text, trans), ...] 该角色的全部台词
    返回 {对方: (日文形式, Counter(中文译法))}
    """
    forms = defaultdict(Counter)
    zh_seen = defaultdict(Counter)
    targets = [t for t in list(cast) + ADDRESS_EXTRA if t != speaker]
    for text, trans in pairs:
        for other in targets:
            if other not in text:
                continue
            for m in re.finditer(re.escape(other), text):
                tail = text[m.end():m.end() + 3]
                suffix = next((h for h in HONORIFICS if tail.startswith(h)), "")
                forms[other][suffix] += 1
                if trans:
                    zh_seen[(other, suffix)][trans] += 1
    return forms, zh_seen


# 中文里敬称的常见落法，按优先级匹配译文中实际出现的形式
ZH_SUFFIX_PATTERNS = ["前辈", "学姐", "学长", "老师", "大人", "酱", "桑", "君", "同学"]


def zh_address(zh_base, trans_samples):
    """从该称呼实际出现过的译文里，反推中文落法。命中不了就退回裸名。"""
    hits = Counter()
    for trans, n in trans_samples.items():
        if zh_base not in trans:
            continue
        for m in re.finditer(re.escape(zh_base), trans):
            tail = trans[m.end():m.end() + 2]
            suffix = next((s for s in ZH_SUFFIX_PATTERNS if tail.startswith(s)), "")
            hits[zh_base + suffix] += n
            break
        if trans[max(0, trans.find(zh_base) - 1):].startswith("小" + zh_base):
            hits["小" + zh_base] += n
    return hits.most_common(1)[0][0] if hits else zh_base


# 称呼表（xlsx）的行列表头是全名，CSV 的 name 列是显示名，需要映射。
# P 不建卡（玩家），但仍要保留"别人怎么称呼他"。
XLSX_NAME_TO_KEY = {
    "花海 咲季": "咲季",
    "月村 手毬": "手毬",
    "藤田 ことね": "ことね",
    "有村 麻央": "麻央",
    "葛城 リーリヤ": "リーリヤ",
    "倉本 千奈": "千奈",
    "紫雲 清夏": "清夏",
    "篠澤 広": "広",
    "姫崎 莉波": "莉波",
    "花海 佑芽": "佑芽",
    "P": "プロデューサー",
    "秦谷 美鈴": "美鈴",
    "十王 星南": "星南",
    "根緒 亜紗里": "あさり先生",
    "真城 優": "優",
    "雨夜 燕": "燕",
    "賀陽 燐羽": "燐羽",
    "白草 月花": "月花",
    "白草 四音": "四音",
    "藍井 撫子": "撫子",
}


def cell_lines(value):
    return [ln.strip() for ln in str(value or "").split("\n") if ln.strip()]


def pair_forms(jp_cell, zh_cell):
    """把一格里的多种称呼按行对齐成 "日文 → 中文" 列表。

    行数对不上说明两个 sheet 不同步——宁可整格跳过，也不要错配出一条
    权威性很高却译错的硬约束。
    """
    jp, zh = cell_lines(jp_cell), cell_lines(zh_cell)
    if not jp or not zh:
        return []
    if len(jp) != len(zh):
        return None
    return [f"{a} → {b}" for a, b in zip(jp, zh)]


def load_xlsx_forms(path):
    """读称呼表 xlsx，返回 {说话人: {"self": 自称, "address": {对方: 称呼}}}。

    行=说话人，列=被称呼者，对角线=自称。sheet1 原文，sheet2 译文。
    """
    try:
        import openpyxl
    except ImportError:
        sys.exit("读 xlsx 需要 openpyxl：pip install openpyxl")

    wb = openpyxl.load_workbook(path)
    if len(wb.worksheets) < 2:
        sys.exit(f"{path} 需要两个 sheet（原版 / 翻译版）")
    jp_ws, zh_ws = wb.worksheets[0], wb.worksheets[1]
    jp_rows = list(jp_ws.iter_rows(values_only=True))
    zh_rows = list(zh_ws.iter_rows(values_only=True))

    header = [str(c).strip() if c else "" for c in jp_rows[0]]
    result, skipped = {}, []
    for jp_row, zh_row in zip(jp_rows[1:], zh_rows[1:]):
        speaker = XLSX_NAME_TO_KEY.get(str(jp_row[0] or "").strip())
        if not speaker:
            continue
        entry = {"self": "", "address": {}}
        for col in range(1, min(len(header), len(jp_row), len(zh_row))):
            target = XLSX_NAME_TO_KEY.get(header[col])
            if not target:
                continue
            forms = pair_forms(jp_row[col], zh_row[col])
            if forms is None:
                skipped.append(
                    f"{speaker}→{target}: 原文 {len(cell_lines(jp_row[col]))} 条"
                    f"{cell_lines(jp_row[col])}，译文 {len(cell_lines(zh_row[col]))} 条"
                    f"{cell_lines(zh_row[col])}")
                continue
            if not forms:
                continue
            if target == speaker:
                # 自称译文与原文相同 = 那一格忘了翻。人称代词不存在原样保留的情况，
                # 写进去就成了"把 私 译作 私"的硬约束，比没有约束更糟——直接丢掉
                good = [f for f, jp in zip(forms, cell_lines(jp_row[col]))
                        if f != f"{jp} → {jp}"]
                if len(good) != len(forms):
                    print(f"!! {speaker} 自称疑似漏翻（译文与原文相同），已丢弃: "
                          f"{[f for f in forms if f not in good]}")
                if good:
                    entry["self"] = "；".join(good)
            else:
                entry["address"][target] = "；".join(forms)
        result[speaker] = entry
    if skipped:
        print(f"!! {len(skipped)} 格原文/译文行数不一致，已跳过（错配会写出一条权威但错的硬约束）：")
        for s in skipped:
            print("   ", s)
    print(f"从称呼表读入 {len(result)} 个角色")
    return result


def apply_xlsx(cards, forms):
    """称呼表是人工整理的，**整张替换**该角色的称呼表，不与统计结果合并。

    留空的格子意味着"没有固定叫法"，就该不给约束——用统计结果去填空会把
    人工表和噪声混在一张卡里，而卡是硬约束，来源必须干净。
    """
    for name, entry in forms.items():
        card = cards.get(name)
        if not card:
            continue  # 不在建卡范围内（如 P），但它作为被称呼者仍然生效
        if entry["self"]:
            card["first_person"] = entry["self"]
        card["address"] = entry["address"]
        card["_stats"]["address_from"] = "称呼表"
    return cards


ZH_FALLBACK = {"プロデューサー": "制作人"}


def zh_of(name, name_dict):
    """中文名：字典里有就用字典，没有说明本身就是汉字名，原样保留。"""
    return name_dict.get(name) or ZH_FALLBACK.get(name, name)


def build(rows, cast, counts, name_dict):
    by_name = defaultdict(list)
    for name, text, trans in rows:
        if name in cast:
            by_name[name].append((text, trans))

    cards = {}
    for name in cast:
        pairs = by_name[name]
        fp = first_person([t for t, _ in pairs])
        forms, zh_seen = address_forms(pairs, cast, name)
        addr = {}
        for other, suffixes in forms.items():
            suffix, n = suffixes.most_common(1)[0]
            if n < 5:  # 偶发提及，不足以定为固定称呼
                continue
            jp = other + suffix
            zh = zh_address(zh_of(other, name_dict), zh_seen[(other, suffix)])
            addr[other] = f"{jp} → {zh}"
        cards[name] = {
            "zh": zh_of(name, name_dict),
            "first_person": " / ".join(f"{t}（{c}次）" for t, c in fp.most_common(3)),
            "politeness": "",
            "speech": [],
            "address": dict(sorted(addr.items(), key=lambda kv: -forms[kv[0]].total())),
            "_stats": {"lines": counts[name]},
        }
    return cards


def merge_existing(cards, path):
    """保留人工已填写的字段，只刷新统计出来的部分。"""
    if not path.exists():
        return cards
    old = json.loads(path.read_text(encoding="utf-8"))
    for name, card in cards.items():
        prev = old.get(name)
        if not prev:
            continue
        for key in ("politeness", "speech"):
            if prev.get(key):
                card[key] = prev[key]
        # 人工改过的称呼条目优先
        card["address"] = {**card["address"], **prev.get("address", {})}
        if prev.get("zh"):
            card["zh"] = prev["zh"]
        if prev.get("first_person") and not prev["first_person"].endswith("次）"):
            card["first_person"] = prev["first_person"]
    for name, prev in old.items():
        cards.setdefault(name, prev)
    return cards


CLAUSE_SPLIT = re.compile(r"(?:\\n|<br>|[。！？!?♪、…]+)")


def speech_stats(rows, cast, top=12):
    """统计每个角色的语气候选：句尾形式 + 特征词。

    用相对全体的倍率（lift）排序而不是绝对频次——「です」谁都在说，
    只有相对别人明显偏高的才是这个角色的口癖。
    输出只是候选，规则要人工挑完再写成可执行的形式。
    """
    by_name = defaultdict(list)
    for name, text, _ in rows:
        if name in cast:
            by_name[name].append(text)

    def features(lines):
        endings, grams = Counter(), Counter()
        for text in lines:
            for clause in CLAUSE_SPLIT.split(text):
                clause = clause.strip()
                if len(clause) < 2:
                    continue
                endings[clause[-2:]] += 1
                if len(clause) >= 3:
                    endings[clause[-3:]] += 1
                for n in (2, 3):
                    for i in range(len(clause) - n + 1):
                        grams[clause[i:i + n]] += 1
        return endings, grams

    all_end, all_gram = Counter(), Counter()
    per = {}
    for name in cast:
        e, g = features(by_name[name])
        per[name] = (e, g)
        all_end.update(e)
        all_gram.update(g)

    def distinctive(mine, total, min_count):
        n_mine, n_all = sum(mine.values()), sum(total.values())
        out = []
        for token, c in mine.items():
            if c < min_count:
                continue
            lift = (c / n_mine) / (total[token] / n_all)
            if lift > 1.5:
                out.append((lift, c, token))
        out.sort(reverse=True)
        return out[:top]

    print("\n" + "=" * 70)
    print("语气候选（lift = 相对全体的偏好倍率，越高越是这个角色专有）")
    print("这是候选不是结论：挑出真正的口癖，再写成可执行的中文处理规则。")
    print("=" * 70)
    for name in cast:
        e, g = per[name]
        print(f"\n【{name}】 {len(by_name[name])} 行")
        ends = distinctive(e, all_end, 15)
        words = distinctive(g, all_gram, 20)
        seen_end = {t for _, _, t in ends}
        print("  句尾：", "  ".join(f"{t}(x{l:.1f},{c})" for l, c, t in ends) or "（无明显偏好）")
        rest = [(l, c, t) for l, c, t in words if t not in seen_end]
        print("  特征词：", "  ".join(f"{t}(x{l:.1f},{c})" for l, c, t in rest[:top]) or "（无）")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--top", type=int, default=16,
                    help="建卡的角色数量（默认 16，覆盖约 95%% 台词）")
    ap.add_argument("--xlsx", default="",
                    help="人工整理的称呼表 xlsx（sheet1 原文 / sheet2 译文），优先级高于统计")
    ap.add_argument("--out", default=str(OUT))
    ap.add_argument("--speech-stats", action="store_true",
                    help="打印各角色的语气候选（句尾/特征词），供人工写 speech 规则")
    args = ap.parse_args()

    name_dict = json.loads(NAME_DICT.read_text(encoding="utf-8")) if NAME_DICT.exists() else {}
    rows = load_rows()
    if not rows:
        sys.exit("csv_data 里没有人工译文，无法统计")
    cast, counts = pick_cast(rows, args.top)
    # 顺序要紧：先并回上一版的人工字段，再让称呼表覆盖——称呼表优先级最高
    cards = merge_existing(build(rows, cast, counts, name_dict), Path(args.out))
    if args.xlsx:
        cards = apply_xlsx(cards, load_xlsx_forms(args.xlsx))

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(cards, ensure_ascii=False, indent=2) + "\n",
                   encoding="utf-8")
    print(f"写入 {len(cards)} 张角色卡 -> {out}")

    if args.speech_stats:
        speech_stats(rows, cast)

    todo = [n for n, c in cards.items() if not c["speech"]]
    if todo:
        print(f"\n!! 以下 {len(todo)} 张卡的 speech（语气规则）还没人工填写：")
        print("  " + "、".join(todo))
        print("  统计只能给出自称和称呼，语气规则必须人工写，且必须可执行。")


if __name__ == "__main__":
    main()
